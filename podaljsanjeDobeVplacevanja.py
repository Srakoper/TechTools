"""
Script for automatic processing of "Podaljšanje dobe vplačevanja" request tickets.
Requires a set of exported tables for a given policy in a .xlsx format. Use SAP/Podatki o polici.sql commands to generate data for export.
Requires manual input of extension period and sign date, which should be found in an enclosed .pdf document from a request ticket.
Generates appropriate UPDATE and INSERT SQL commands, as well as a summary of the update process.
Instructions for use:
    1. export tables for a given policy and save .xlsx file in the directory containing this script
    3. run podaljsanjeDobeVplacevanja.py
    4. enter file name of .xlsx file, extension period, amd sign date
    5. generated data is saved to podaljsanje_dobe.txt file in the directory containing this script
"""
from os import getcwd
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta

def verifyDate(date_string):
    """
    Verifies input date in D.M.YYYY format.
    :param date_string: str; string of date in D.M.YYYY format
    :return: True if valid date, False otherwise
    """
    if len(date_string) < 8 or len(date_string) > 10: return False
    for i in range(len(date_string)):
        if date_string[1] == ".":
            if int(date_string[0]) <= 0: return False
            if date_string[3] == ".":
                if int(date_string[2]) <= 0: return False
                if int(date_string[4:]) <= 999: return False
            else:
                if int(date_string[2:4]) <= 0 or int(date_string[2:4]) > 12: return False
                if int(date_string[5:]) <= 999: return False
        else:
            if int(date_string[:2]) <= 0 or int(date_string[:2]) > 31: return False
            if date_string[4] == ".":
                if int(date_string[3]) <= 0: return False
                if int(date_string[5:]) <= 999: return False
            else:
                if int(date_string[3:5]) <= 0 or int(date_string[3:5]) > 12: return False
                if int(date_string[6:]) <= 999: return False
    return True

def getWorkbook(filename, dir):
    """
    Fetches Excel Workbook object from a given .xlsx filename.
    :param filename: str; .xlsx filename
    :param dir: str; current working directory path
    :return: Workbook object
    """
    return load_workbook(dir + filename, data_only = True)

def getData(workbook):
    """
    Fetches relevant data from the Workbook object.
    :param workbook: Workbook object
    :return: dict; dictionary of relevant data
    """
    sheet1 = workbook["Select policy"]
    sheet2 = workbook["Select pol_benf"]
    policy_policy_no           = int (sheet1["C2"].value)
    policy_pol_ref_no          = int(sheet1["D2"].value)
    policy_end_date            = sheet1["I2"].value
    policy_term                = int(sheet1["J2"].value)
    policy_payout_start_date   = sheet1["S2"].value
    pol_benf_term              = int(sheet2["Q2"].value)
    pol_benf_payment_term      = int(sheet2["R2"].value)
    pol_benf_end_date          = sheet2["V2"].value
    pol_benf_prem_stop_date    = sheet2["W2"].value
    pol_benf_prem_stop_date_17 = None
    for cell in reversed(sheet2["W"]):
        if cell:
            pol_benf_prem_stop_date_17 = cell.value
            break
    warning = None
    try: assert str(pol_benf_prem_stop_date - pol_benf_prem_stop_date_17) == "1 day, 0:00:00"
    except AssertionError: warning = "PREM_STOP_DATE z BENFNO = 17 v pol_benf ni za 1 dan manjši od ostalih PREM_STOP_DATE."
    return {"warning": warning,
            "policy_policy_no": policy_policy_no,
            "policy_pol_ref_no": policy_pol_ref_no,
            "policy_end_date": policy_end_date,
            "policy_term": policy_term,
            "policy_payout_start_date": policy_payout_start_date,
            "pol_benf_term": pol_benf_term,
            "pol_benf_payment_term": pol_benf_payment_term,
            "pol_benf_end_date": pol_benf_end_date,
            "pol_benf_prem_stop_date": pol_benf_prem_stop_date,
            "pol_benf_prem_stop_date_17": pol_benf_prem_stop_date_17}

def generateOutput(data, y, m, sign_date):
    """
    Generates output data for 'Podaljšanje dobe varčevanja' from given input.
    :param data: dict; input data
    :param y: int; years to extend
    :param m: int; months to extend
    :param sign_date: str; sign date in D.M.YYYY format
    :return: str; generated output data
    """
    policy_end_date_new                = data["policy_end_date"] + relativedelta(years=y) + relativedelta(months=m)
    policy_payout_start_date_new       = data["policy_payout_start_date"] + relativedelta(years=y) + relativedelta(months=m)
    pol_benf_end_date_new              = data["pol_benf_end_date"] + relativedelta(years=y) + relativedelta(months=m)
    pol_benf_prem_stop_date_new        = data["pol_benf_prem_stop_date"] + relativedelta(years=y) + relativedelta(months=m)
    pol_benf_prem_stop_date_17_new     = data["pol_benf_prem_stop_date_17"] + relativedelta(years=y) + relativedelta(months=m)
    str_policy_pol_ref_no              = data["policy_pol_ref_no"]
    str_policy_end_date_new            = policy_end_date_new.strftime("%#d.%#m.%Y")
    str_policy_payout_start_date_new   = policy_payout_start_date_new.strftime("%#d.%#m.%Y")
    str_policy_term                    = data["policy_term"]
    str_policy_term_new                = data["policy_term"] + (y * 12 + m)
    str_pol_benf_payment_term_new      = data["pol_benf_payment_term"] + (y * 12 + m)
    str_pol_benf_term_new              = data["pol_benf_term"] + (y * 12 + m)
    str_pol_benf_end_date_new          = pol_benf_end_date_new.strftime("%#d.%#m.%Y")
    str_pol_benf_prem_stop_date_17_new = pol_benf_prem_stop_date_17_new.strftime("%#d.%#m.%Y")
    str_pol_benf_prem_stop_date_new    = pol_benf_prem_stop_date_new.strftime("%#d.%#m.%Y")
    description         = "Popraviti v tabeli policy za POL_REF_NO = {}\nEND_DATE = {}\nPAYOUT_START_DATE = {}\nTERM = {}\nNEXT_PAYOUT_DATE: če je notri, naj se datum briše\n\nPopraviti v tabeli pol_benf za POL_REF_NO = {}:\nPAYMENT_TERM = {}\nTERM = {}\nEND_DATE = {}\nPREM_STOP_DATE = {} za BENFNO = 17 / {} za vse ostale BENFNO\n\nVstaviti v tabelo pol_endosrements zapis o podaljšanju:\nENDORSE_TYPE = 47\nPOL_REF_NO = {}\nCHANGE_DESC = Podaljšanje dobe varčevanja\nOLD_VALUE = {} (sprememba term zaradi podaljšanja dobe varčevanja)\nNEW_VALUE = {}\nEFFECTIVE_DATE = {} (datum podpisa, če z ePeresom, drugače datum prejema obrazca)\nTRANSACTION_DATE = <trenutni_datum>"\
        .format(str_policy_pol_ref_no,
                str_policy_end_date_new,
                str_policy_payout_start_date_new,
                str_policy_term_new,
                str_policy_pol_ref_no,
                str_pol_benf_payment_term_new,
                str_pol_benf_term_new,
                str_pol_benf_end_date_new,
                str_pol_benf_prem_stop_date_17_new,
                str_pol_benf_prem_stop_date_new,
                str_policy_pol_ref_no,
                str_policy_term,
                str_policy_term_new,
                sign_date)
    SQL_policy          = "UPDATE policy\n   SET END_DATE = '{}',\n       PAYOUT_START_DATE = '{}',\n       TERM = {},\n       NEXT_PAYOUT_DATE = NULL\n WHERE POL_REF_NO = {};"\
        .format(str_policy_end_date_new,
                str_policy_payout_start_date_new,
                str_policy_term_new,
                str_policy_pol_ref_no)
    SQL_pol_benf        = "UPDATE pol_benf\n   SET PAYMENT_TERM = {},\n       TERM = {},\n       END_DATE = '{}',\n       PREM_STOP_DATE = '{}'\n WHERE POL_REF_NO = {}\n   AND BENFNO <> 17;"\
        .format(str_pol_benf_payment_term_new,
                str_pol_benf_term_new,
                str_pol_benf_end_date_new,
                str_pol_benf_prem_stop_date_new,
                str_policy_pol_ref_no)
    SQL_pol_benf_17     = "UPDATE pol_benf\n   SET PAYMENT_TERM = {},\n       TERM = {},\n       END_DATE = '{}',\n       PREM_STOP_DATE = '{}'\n WHERE POL_REF_NO = {}\n   AND BENFNO = 17;"\
        .format(str_pol_benf_payment_term_new,
                str_pol_benf_term_new,
                str_pol_benf_end_date_new,
                str_pol_benf_prem_stop_date_17_new,
                str_policy_pol_ref_no)
    SQL_pol_endorsments = "INSERT INTO pol_endorsements\n  (SITENO,\n   ENDORSE_TYPE,\n   POL_REF_NO,\n   CHANGE_DESC,\n   OLD_VALUE,\n   NEW_VALUE,\n   EFFECTIVE_DATE,\n   TRANSACTION_DATE,\n   LETTER_SENT,\n   STATUS)\nVALUES\n  ({},\n   {},\n   {},\n   '{}',\n   {},\n   {},\n   '{}',\n   {},\n   '{}',\n   '{}');"\
        .format(7,
                47,
                str_policy_pol_ref_no,
                "Podaljšanje dobe varčevanja",
                str_policy_term,
                str_policy_term_new,
                sign_date,
                "sysdate",
                "N",
                "A")
    return SQL_policy + "\n\n" + SQL_pol_benf  + "\n\n" + SQL_pol_benf_17 + "\n\n" + SQL_pol_endorsments + '\n\nPovzetek popravkov iz zgornjih SQL ukazov:\n\n' + description + "\n\nPriložen izvoz trenutnega stanja tabel za polico {}.".format(data["policy_policy_no"])

def main():
    path = getcwd().replace("\\", "\\\\") + "\\\\"
    while True:
        try:
            wb = getWorkbook(input("Name of .xlsx file with policy data: "), path)
            break
        except FileNotFoundError: print("File not found. Please try again.")
    while True:
        try:
            extension = int(input("Extension period in months: "))
            break
        except: print("Enter valid extension period.")
    years = extension // 12
    months = extension % 12
    while True:
        sign_date = input("Sign date in D.M.YYYY format: ")
        if verifyDate(sign_date): break
        else: print("Enter date in valid D.M.YYYY format.")
    fh = open("podaljsanje_dobe.txt", "w")
    fh.write(generateOutput(getData(wb), years, months, sign_date))
    fh.close()
    print("Output data saved to file podaljsanje_dobe.txt.")

if __name__ == "__main__":
    main()