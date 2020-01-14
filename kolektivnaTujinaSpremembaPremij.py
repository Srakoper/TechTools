from os import getcwd
from openpyxl import load_workbook
from datetime import date

class Policy(object):
    def __init__(self,
                 policy_no,
                 pol_ref_no,
                 start_date,
                 month,
                 effective_date,
                 duration,
                 cancelled_premium,
                 new_gross_premium):
        self.policy_no         = policy_no
        self.pol_ref_no        = pol_ref_no
        self.start_date        = start_date
        self.month             = month
        self.effective_date    = effective_date
        self.duration          = duration
        self.cancelled_premium = cancelled_premium
        self.new_gross_premium = new_gross_premium
    def setMonth(self, date): self.month = date.month
    def setEffectiveDate(self, current, extracted):
        """
        If current month equals extracted month, no effective date is set.
        """
        if current[0] > extracted: self.effective_date = date(current[1], extracted, 10)
        elif current[0] < extracted: self.effective_date = date(current[1] - 1, extracted, 10)
    def getPolicy_no(self):         return self.policy_no
    def getPol_ref_no(self):        return self.pol_ref_no
    def getStart_date(self):        return self.start_date
    def getDuration(self):          return self.duration
    def getCancelled_premium(self): return self.cancelled_premium
    def getNew_gross_premium(self): return self.new_gross_premium
    def getMonth(self):             return self.month
    def getEffective_date(self):    return self.effective_date
    def __repr__(self):             return "policy_no: {}\npol_ref_no: {}\nstart_date: {}\nmonth: {}\neffective_date: {}\nduration: {}\ncancelled_premium: {}\nnew_gross_premium {}" \
                                           .format(self.policy_no,
                                                   self.pol_ref_no,
                                                   self.start_date,
                                                   self.month,
                                                   self.effective_date,
                                                   self.duration,
                                                   self.cancelled_premium,
                                                   self.new_gross_premium)

def getWorkbook(filename, dir):
    """
    Fetches Excel Workbook object from a given .xlsx filename.
    :param filename: str; .xlsx filename
    :param dir: str; current working directory path
    :return: Workbook object
    """
    return load_workbook(dir + filename, data_only = True)

def getData(workbook):
    """
    Fetches relevant data from the Workbook object.
    :param workbook: Workbook object
    :return: tuple; tuple of relevant data
    """
    data = list()
    sheet = workbook["Select select"]
    assert sheet["B"][0].value == "POLICY_NO" \
       and sheet["C"][0].value == "POL_REF_NO" \
       and sheet["F"][0].value == "START_DATE" \
       and sheet["G"][0].value == "TRAJANJE" \
       and sheet["H"][0].value == "STORNO_LETNA_PREMIJA" \
       and sheet["J"][0].value == "NOVA_PREMIJA_BRUTO"
    for i in range(1, len(sheet["B"])):
        if sheet["B"][i].value:
            data.append(Policy(sheet["B"][i].value,
                               round(sheet["C"][i].value),
                               sheet["F"][i].value,
                               None,
                               None,
                               sheet["G"][i].value,
                               sheet["H"][i].value,
                               sheet["J"][i].value))
    return tuple(data)

def generateOutput(policies_data, current_dates):
    """
    :param data: tuple; tuple of policy data
    :return: string; generated SQL statements for given policy data
    """
    output_policy           = list()
    output_pol_benf         = list()
    output_pol_endorsements = list()
    for policy in policies_data:
        policy.setMonth(policy.getStart_date())
        policy.setEffectiveDate(current_dates, policy.getMonth())
        output_policy.append("UPDATE policy SET BASIC_PREMIUM = {}, GROSS_PREMIUM = {} WHERE POL_REF_NO = {};" \
                             .format(policy.getNew_gross_premium(),
                                     policy.getNew_gross_premium(),
                                     policy.getPol_ref_no()))
        if policy.getEffective_date(): # checks if effective_date is set
            output_pol_benf.append("UPDATE pol_benf SET PREMIUM = {} WHERE POL_REF_NO = {};" \
                                   .format(policy.getNew_gross_premium(),
                                           policy.getPol_ref_no()))
            output_pol_endorsements.append("INSERT INTO pol_endorsements (SITENO, ENDORSE_TYPE, POL_REF_NO, BENFNO, CHANGE_DESC, BENF_ORD, OLD_VALUE, NEW_VALUE, EFFECTIVE_DATE, TRANSACTION_DATE) VALUES (7, 7, {}, 80, 'Change of Premium', 1, {}, {}, '{}', sysdate);" \
                                           .format(policy.getPol_ref_no(),
                                                   policy.getCancelled_premium(),
                                                   policy.getNew_gross_premium(),
                                                   policy.getEffective_date().strftime("%d.%m.%Y")))
    return "{}\n\n{}\n\n{}" \
           .format("\n".join(output_policy),
                   "\n".join(output_pol_benf),
                   "\n".join(output_pol_endorsements))

def main():
    path = getcwd().replace("\\", "\\\\") + "\\\\"
    current_month_year = (date.today().month, date.today().year)
    while True:
        filename_input = input("Ime .xlsx datoteke s tabelami o polici (X + ENTER za izhod): ")
        if filename_input in ("x", "X"): quit()
        try:
            if filename_input.find(".") == -1: filename_input += ".xlsx"
            wb = getWorkbook(filename_input, path)
            break
        except FileNotFoundError: print("Datoteka ne obstaja. Poskusi znova.")
    policies = getData(wb)
    fh = open("kolektivna_tujina_sprememba_premij_{}_{}.txt".format(current_month_year[1], "0" + str(current_month_year[0]) if len(str(current_month_year[0])) == 1 else current_month_year[0]), "w")
    fh.write("{}".format(generateOutput(policies, current_month_year)))
    fh.close()
    print("Ustvarjeni podatki shranjeni v datoteko kolektivna_tujina_sprememba_premij_{}_{}.txt".format(current_month_year[1], "0" + str(current_month_year[0]) if len(str(current_month_year[0])) == 1 else current_month_year[0]))

if __name__ == "__main__":
    main()