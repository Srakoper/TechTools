import openpyxl

def saveToXLSX(path:str, filename:str, data:tuple, sheetnames:tuple=None) -> str:
    if not path.endswith("\\"): path += "\\"
    if sheetnames:
        try: assert len(data) == len(sheetnames)
        except AssertionError:
            print("DATA and SHEETNAMES tuples must be of same size.")
            exit()
    if not filename.endswith(".xlsx"): filename += ".xlsx"
    wb = openpyxl.Workbook()
    if sheetnames:
        sheet = wb.active
        sheet.title = sheetnames[0]
        for i in range(len(data) - 1):
            wb.create_sheet(title=sheetnames[i + 1])
    for i in range(len(data)):
        sheet = wb.worksheets[i]
        for ii in range(len(data[i][0])): # sets header from SQL results data tuple
            sheet[f"{openpyxl.utils.cell.get_column_letter(ii + 1)}{1}"] = data[i][0][ii]
        for ii in range(len(data[i][1])):
            for iii in range(len(data[i][1][ii])): # sets rows of data from SQL results data tuple
                sheet[f"{openpyxl.utils.cell.get_column_letter(iii + 1)}{ii + 2}"] = data[i][1][ii][iii]
    wb.save(path + filename)
    return f"Data saved successfully to {path + filename}."