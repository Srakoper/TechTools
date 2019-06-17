"""
Script for automatic processing of "Popravek izplačil" request tickets.
Requires a set of exported tables for a given policy in a .xlsx format. Use SAP/Podatki o polici.sql commands to generate data for export.
Generates appropriate UPDATE and DELETE SQL commands, as well as a summary of the update process.
Instructions for use:
    1. export tables for a given policy and save .xlsx file in the directory containing this script
    3. run popravekIzplacil.py
    4. enter file name of .xlsx file
    5. generated data is saved to popravek_izplacil.txt file in the directory containing this script
"""
from os import getcwd
from openpyxl import load_workbook

def getWorkbook(filename, dir):
    """
    Fetches Excel Workbook object from a given .xlsx filename.
    :param filename: str; .xlsx filename
    :param dir: str; current working directory path
    :return: Workbook object
    """
    return load_workbook(dir + filename, data_only = True)

def getData(workbook):
    """
    Fetches relevant data from the Workbook object.
    :param workbook: Workbook object
    :return: dict; dictionary of relevant data
    """
    sheet1 = workbook["Select policy"]
    sheet2 = workbook["Select claims_payment"]
    sheet3 = workbook["Select sa_postings"]
    sheet4 = workbook["Select claim_dets"]
    policy_policy_no                    = int(sheet1["C2"].value)
    claims_payment_claims_payment_id_1  = int(sheet2["C2"].value)
    claims_payment_claims_payment_id_2  = int(sheet2["C3"].value)
    claims_payment_amount_1             = float(sheet2["J2"].value)
    claims_payment_amount_2             = float(sheet2["J3"].value)
    claims_payment_curr_amount_1        = float(sheet2["L2"].value)
    claims_payment_curr_amount_2        = float(sheet2["L3"].value)
    sa_postings_personal_income_tax_ids = list()
    for i in range(len(sheet3["B"])):
        if sheet3["B"][i].value == "Maturity Claim Payment (To Client)" or sheet3["B"][i].value == "Surrender Claim Payment (To Client)":
            if float(sheet3["Q"][i].value) == claims_payment_amount_2:
                sa_postings_maturity_claim_payment_id = int(sheet3["D"][i].value)
                sa_postings_sa_transno = int(sheet3["E"][i].value)
            else:
                sa_postings_maturity_claim_payment_id = None
                sa_postings_sa_transno = None
        if sheet3["B"][i].value == "Personal Income tax (Maturity)" or sheet3["B"][i].value == "Personal Income tax (Surrender)":
            if float(sheet3["P"][i].value) == claims_payment_amount_1 or float(sheet3["Q"][i].value) == claims_payment_amount_1: sa_postings_personal_income_tax_ids.append(sheet3["D"][i].value)
    for i in range(0, len(sheet4["F"])):
        try:
            if int(sheet4["G"][i].value) == 6 and "{:.2f}".format(sheet4["F"][i].value) != "{:.2f}".format(claims_payment_amount_1 + claims_payment_amount_2): claims_dets_claim_amt = True
        except ValueError: claims_dets_claim_amt = False
    return {"policy_policy_no": policy_policy_no,
            "claims_payment_claims_payment_id_1": claims_payment_claims_payment_id_1,
            "claims_payment_claims_payment_id_2": claims_payment_claims_payment_id_2,
            "claims_payment_amount_1": claims_payment_amount_1,
            "claims_payment_amount_2": claims_payment_amount_2,
            "claims_payment_curr_amount_1": claims_payment_curr_amount_1,
            "claims_payment_curr_amount_2": claims_payment_curr_amount_2,
            "sa_postings_maturity_claim_payment_id": sa_postings_maturity_claim_payment_id,
            "sa_postings_personal_income_tax_ids": tuple(sa_postings_personal_income_tax_ids),
            "sa_postings_sa_transno": sa_postings_sa_transno,
            "claims_dets_claim_amt": claims_dets_claim_amt}

def generateOutput(data):
    """
    Generates output data for 'Popravek izplačil' from given input.
    :param data: dict; input data
    :return: str; generated output data
    """
    claims_payment_amount_new      = data["claims_payment_amount_1"] + data["claims_payment_amount_2"]
    claims_payment_curr_amount_new = data["claims_payment_curr_amount_1"] + data["claims_payment_curr_amount_2"]
    description1              = "V tabeli claims_payment, kjer je CLAIMS_PAYMENT_ID = {}, je treba popraviti vrednosti v stolpcih AMOUNT in CURR_AMOUNT:\n\nAMOUNT stara vrednost: {}\nAMOUNT nova vrednost: {}\n\nCURR_AMOUNT stara vrednost: {}\nCURR_AMOUNT nova vrednost: {}\n\nV isti tabeli je treba odstraniti zapis, kjer je CLAIMS_PAYMENT_ID = {}."\
        .format(data["claims_payment_claims_payment_id_2"],
                "{:.2f}".format(data["claims_payment_amount_2"]).replace(".", ","),
                "{:.2f}".format(claims_payment_amount_new).replace(".", ","),
                "{:.2f}".format(data["claims_payment_curr_amount_2"]).replace(".", ","),
                "{:.2f}".format(claims_payment_curr_amount_new).replace(".", ","),
                data["claims_payment_claims_payment_id_1"])
    description2              = "V tabeli sa_postings, kjer je SA_POSTINGNO = {}, je treba popraviti vrednost v stolpcu SA_POSTINGDR:\n\nstara vrednost: {}\nnova vrednost: {}\n\nV isti tabeli je treba odstraniti zapisa, kjer sta SA_POSTINGNO IN {}."\
        .format(data["sa_postings_maturity_claim_payment_id"],
                "{:.2f}".format(data["claims_payment_amount_2"]).replace(".", ","),
                "{:.2f}".format(claims_payment_amount_new).replace(".", ","),
                data["sa_postings_personal_income_tax_ids"])
    SQL_claims_payment_update = "UPDATE claims_payment\n   SET AMOUNT = {},\n       CURR_AMOUNT = {}\n WHERE CLAIMS_PAYMENT_ID = {};"\
        .format("{:.2f}".format(claims_payment_amount_new),
                "{:.2f}".format(claims_payment_curr_amount_new),
                data["claims_payment_claims_payment_id_2"])
    SQL_claims_payment_delete = "DELETE\n  FROM claims_payment\n WHERE CLAIMS_PAYMENT_ID = {};"\
        .format(data["claims_payment_claims_payment_id_1"])
    SQL_sa_posting_update     = "UPDATE sa_postings\n   SET SA_POSTINGDR = {}\n WHERE SA_POSTINGNO = {};"\
        .format("{:.2f}".format(claims_payment_amount_new),
                data["sa_postings_maturity_claim_payment_id"])
    SQL_sa_posting_delete     = "DELETE\n  FROM sa_postings\n WHERE SA_POSTINGNO IN {};"\
        .format(data["sa_postings_personal_income_tax_ids"])
    if data["claims_dets_claim_amt"] and data["sa_postings_sa_transno"]:
        description3          = "V tabeli claim_dets, kjer je SA_POSTINGNO = {}, je treba popraviti vrednost v stolpcu CLAIM_AMT:\n\nstara vrednost: {}\nnova vrednost: {}"\
            .format(data["sa_postings_sa_transno"],
                    "{:.2f}".format(data["claims_payment_amount_2"]).replace(".", ","),
                    "{:.2f}".format(claims_payment_amount_new).replace(".", ","))
        SQL_claim_dets_update = "UPDATE claim_dets\n   SET CLAIM_AMT = {}\n WHERE SA_POSTINGNO = {};"\
            .format("{:.2f}".format(claims_payment_amount_new),
                    data["sa_postings_sa_transno"])
    else:
        description3 = False
        SQL_claim_dets_update = False
    return "<pre>\n<code class=\"sql\">\n" + SQL_claims_payment_update + "\n\n" + SQL_claims_payment_delete + "\n\n" + SQL_sa_posting_update + "\n\n" + SQL_sa_posting_delete + ("\n\n" + SQL_claim_dets_update if SQL_claim_dets_update else "") + "\n</code>\n</pre>" + "\n\nPovzetek popravkov iz zgornjih SQL ukazov:\n\n" + description1 + "\n\n" + description2 + ("\n\n" + description3 if description3 else "") + "\n\nPriložen izvoz trenutnega stanja tabel za polico {}.".format(data["policy_policy_no"])

def main():
    path = getcwd().replace("\\", "\\\\") + "\\\\"
    while True:
        filename_input = input("Ime .xlsx datoteke s tabelami o polici (X + ENTER za izhod): ")
        if filename_input in ("x", "X"): quit()
        try:
            if filename_input.find(".") == -1: filename_input += ".xlsx"
            wb = getWorkbook(filename_input, path)
            break
        except FileNotFoundError: print("Datoteka ne obstaja. Poskusi znova.")
    fh = open("popravek_izplacil.txt", "w")
    fh.write(generateOutput(getData(wb)))
    fh.close()
    print("Ustvarjeni podatki shranjeni v datoteko popravek_izplacil.txt.")

if __name__ == "__main__":
    main()