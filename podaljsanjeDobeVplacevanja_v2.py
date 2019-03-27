"""
Script for automatic processing of "Podaljšanje dobe vplačevanja" request tickets.
Requires a set of exported tables for a given policy in a .xlsx format. Use SAP/Podatki o polici.sql commands to generate data for export.
Requires manual input of extension period and sign date, which should be found in an enclosed .pdf document from a request ticket.
Generates appropriate UPDATE, INSERT and DELETE SQL commands.
Instructions for use:
    1. export tables for a given policy and save .xlsx file in the directory containing this script
    3. run podaljsanjeDobeVplacevanja.py
    4. enter file name of .xlsx file, extension period, and sign date
    5. generated data is saved to podaljsanje_dobe.txt file in the directory containing this script
"""
from os import getcwd
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta

def verifyDate(date_string):
    """
    Verifies input date in D.M.YYYY format.
    :param date_string: str; string of date in D.M.YYYY format
    :return: True if valid date, False otherwise
    """
    if len(date_string) < 8 or len(date_string) > 10: return False
    for i in range(len(date_string)):
        if date_string[1] == ".":
            if int(date_string[0]) <= 0: return False
            if date_string[3] == ".":
                if int(date_string[2]) <= 0: return False
                if int(date_string[4:]) <= 999: return False
            else:
                if int(date_string[2:4]) <= 0 or int(date_string[2:4]) > 12: return False
                if int(date_string[5:]) <= 999: return False
        else:
            if int(date_string[:2]) <= 0 or int(date_string[:2]) > 31: return False
            if date_string[4] == ".":
                if int(date_string[3]) <= 0: return False
                if int(date_string[5:]) <= 999: return False
            else:
                if int(date_string[3:5]) <= 0 or int(date_string[3:5]) > 12: return False
                if int(date_string[6:]) <= 999: return False
    return True

def convertDateDMYYYY(date):
    """
    Converts date from YYYY-MM-DD format to D.M.YYYY format.
    :param date: str; date in YYYY-MM-DD format
    :return: str; converted date in D.M.YYYY format
    """
    (year, month, day) = date.split("-")
    return "{}.{}.{}".format(int(day), int(month), int(year))

def getWorkbook(filename, dir):
    """
    Fetches Excel Workbook object from a given .xlsx filename.
    :param filename: str; .xlsx filename
    :param dir: str; current working directory path
    :return: Workbook object
    """
    return load_workbook(dir + filename, data_only = True)

def getData(workbook):
    """
    Fetches relevant data from the Workbook object.
    :param workbook: Workbook object
    :return: dict; dictionary of relevant data
    """
    sheet1 = workbook["Select policy"]
    sheet2 = workbook["Select pol_benf"]
    policy_policy_no                 = int(sheet1["C2"].value)
    policy_pol_ref_no                = int(sheet1["D2"].value)
    policy_end_date                  = sheet1["I2"].value
    policy_term                      = int(sheet1["J2"].value)
    policy_payout_start_date         = sheet1["S2"].value
    policy_payout_start_date_minus_1 = policy_payout_start_date - relativedelta(days=1) if policy_payout_start_date else None
    policy_next_payout_date          = sheet1["T2"].value
    pol_benf_term                    = int(sheet2["Q2"].value)
    pol_benf_payment_term            = int(sheet2["R2"].value)
    pol_benf_end_date                = sheet2["V2"].value
    pol_benf_prem_stop_date          = sheet2["W2"].value
    pol_benf_prem_stop_date_17       = None
    for cell in reversed(sheet2["W"]):
        if cell:
            pol_benf_prem_stop_date_17 = cell.value
            break
    warning = None
    try: assert str(pol_benf_prem_stop_date - pol_benf_prem_stop_date_17) == "1 day, 0:00:00"
    except AssertionError: warning = "PREM_STOP_DATE z BENFNO = 17 v pol_benf ni za 1 dan manjši od ostalih PREM_STOP_DATE."
    if policy_next_payout_date and workbook["Select scholar_rep"]["B2"].value:
        scholar_rep = True
        claim_pers_dets_claim_id = int(workbook["Select claim_pers_dets"]["B2"].value)
        sa_postings = workbook["Select sa_postings"]
        sa_postings_sa_postingno = list()
        for i in range(1, len(sa_postings["F"])):
            if sa_postings["F"][i].value in [103, 105]: sa_postings_sa_postingno.append(int(sa_postings["D"][i].value))
        claim_dets = workbook["Select claim_dets"]
        claim_dets_claims = dict()
        for i in range(1, len(claim_dets["G"])):
            if claim_dets["G"][i].value == 1: claim_dets_claims["claim_stage_1"] = {"claim_id"   : claim_dets["D"][i].value,
                                                                                    "claim_amt"  : claim_dets["F"][i].value,
                                                                                    "benfno"     : claim_dets["K"][i].value}
            if claim_dets["G"][i].value == 5: claim_dets_claims["claim_stage_5"] = {"claim_id"   : claim_dets["D"][i].value,
                                                                                    "claim_amt"  : claim_dets["F"][i].value,
                                                                                    "benfno"     : claim_dets["K"][i].value}
        ua_holdings = workbook["Select ua_holdings"]
        ua_holdings_funds = set()
        ua_holdings_all = dict()
        for i in range(1, len(ua_holdings["O"])):
            if ua_holdings["L"][i].value in (173, 175): ua_holdings_funds.add(ua_holdings["O"][i].value)
        for fund in ua_holdings_funds:
            temp = dict()
            for i in range(1, len(ua_holdings["O"])):
                if ua_holdings["L"][i].value in (173, 175) and ua_holdings["O"][i].value == fund:
                    temp[ua_holdings["L"][i].value] = {"planno"         : ua_holdings["B"][i].value,
                                                       "start_date"     : ua_holdings["E"][i].value,
                                                       "pol_stat"       : ua_holdings["F"][i].value,
                                                       "due_date"       : ua_holdings["G"][i].value,
                                                       "eff_date"       : ua_holdings["H"][i].value,
                                                       "benfno"         : ua_holdings["I"][i].value,
                                                       "freq"           : ua_holdings["J"][i].value,
                                                       "basic_prem"     : ua_holdings["Q"][i].value,
                                                       "total_prem"     : ua_holdings["R"][i].value,
                                                       "net_prem"       : ua_holdings["T"][i].value,
                                                       "unit_prem"      : ua_holdings["U"][i].value,
                                                       "unit_price"     : ua_holdings["V"][i].value,
                                                       "units_allocated": ua_holdings["W"][i].value,
                                                       "ua_transno"     : ua_holdings["Z"][i].value}
                    ua_holdings_all[fund] = temp
        pol_endorsements = workbook["Select pol_endorsements"]
        benfnos     = list()
        benfnos_all = list()
        for i in range(1, len(pol_endorsements["P"])):
            if pol_endorsements["P"][i].value and pol_endorsements["P"][i].value.strftime("%#d.%#m.%Y") == policy_payout_start_date_minus_1.strftime("%#d.%#m.%Y"): benfnos_all.append(pol_endorsements["K"][i].value)
        for i in range(1, len(sheet2["I"])):
            if sheet2["I"][i].value in benfnos_all and sheet2["N"][i].value == "L": benfnos.append(sheet2["I"][i].value)
        if benfnos: benfnos = tuple(benfnos)
        else: benfnos = None
    else:
        scholar_rep = False
        claim_pers_dets_claim_id = None
        sa_postings_sa_postingno = None
        claim_dets_claims = None
        ua_holdings_all = None
        benfnos = None
    return {"warning"                          : warning,
            "policy_policy_no"                 : policy_policy_no,
            "policy_pol_ref_no"                : policy_pol_ref_no,
            "policy_end_date"                  : policy_end_date,
            "policy_term"                      : policy_term,
            "policy_payout_start_date"         : policy_payout_start_date,
            "policy_payout_start_date_minus_1" : policy_payout_start_date_minus_1,
            "policy_next_payout_date"          : policy_next_payout_date,
            "pol_benf_term"                    : pol_benf_term,
            "pol_benf_payment_term"            : pol_benf_payment_term,
            "pol_benf_end_date"                : pol_benf_end_date,
            "pol_benf_prem_stop_date"          : pol_benf_prem_stop_date,
            "pol_benf_prem_stop_date_17"       : pol_benf_prem_stop_date_17,
            "scholar_rep"                      : scholar_rep,
            "claim_pers_dets_claim_id"         : claim_pers_dets_claim_id,
            "sa_postings_sa_postingno"         : sa_postings_sa_postingno,
            "claim_dets_claims"                : claim_dets_claims,
            "ua_holdings_all"                  : ua_holdings_all,
            "benfnos"                          : benfnos}

def generateOutput(data, y, m, sign_date):
    """
    Generates output data for 'Podaljšanje dobe varčevanja' from given input.
    :param data: dict; input data
    :param y: int; years to extend
    :param m: int; months to extend
    :param sign_date: str; sign date in D.M.YYYY format
    :return: str; generated output data
    """
    policy_end_date_new                  = data["policy_end_date"] + relativedelta(years=y) + relativedelta(months=m)
    policy_payout_start_date_new         = data["policy_payout_start_date"] + relativedelta(years=y) + relativedelta(months=m)
    pol_benf_end_date_new                = data["pol_benf_end_date"] + relativedelta(years=y) + relativedelta(months=m)
    pol_benf_prem_stop_date_new          = data["pol_benf_prem_stop_date"] + relativedelta(years=y) + relativedelta(months=m)
    pol_benf_prem_stop_date_17_new       = data["pol_benf_prem_stop_date_17"] + relativedelta(years=y) + relativedelta(months=m)
    str_policy_policy_no                 = data["policy_policy_no"]
    str_policy_pol_ref_no                = data["policy_pol_ref_no"]
    str_policy_end_date_new              = policy_end_date_new.strftime("%#d.%#m.%Y")
    str_policy_payout_start_date_new     = policy_payout_start_date_new.strftime("%#d.%#m.%Y")
    str_policy_payout_start_date_minus_1 = data["policy_payout_start_date_minus_1"].strftime("%#d.%#m.%Y")
    str_policy_term                      = data["policy_term"]
    str_policy_term_new                  = data["policy_term"] + (y * 12 + m)
    str_pol_benf_payment_term_new        = data["pol_benf_payment_term"] + (y * 12 + m)
    str_pol_benf_term_new                = data["pol_benf_term"] + (y * 12 + m)
    str_pol_benf_end_date_new            = pol_benf_end_date_new.strftime("%#d.%#m.%Y")
    str_pol_benf_prem_stop_date_17_new   = pol_benf_prem_stop_date_17_new.strftime("%#d.%#m.%Y")
    str_pol_benf_prem_stop_date_new      = pol_benf_prem_stop_date_new.strftime("%#d.%#m.%Y")
    SQL_policy          = "UPDATE policy\n   SET END_DATE = '{}',\n       PAYOUT_START_DATE = '{}',\n       TERM = {},\n       NEXT_PAYOUT_DATE = NULL\n WHERE POL_REF_NO = {};"\
        .format(str_policy_end_date_new,
                str_policy_payout_start_date_new,
                str_policy_term_new,
                str_policy_pol_ref_no)
    SQL_pol_benf        = "UPDATE pol_benf\n   SET PAYMENT_TERM = {},\n       TERM = {},\n       END_DATE = '{}',\n       PREM_STOP_DATE = '{}'\n WHERE POL_REF_NO = {}\n   AND BENFNO <> 17;"\
        .format(str_pol_benf_payment_term_new,
                str_pol_benf_term_new,
                str_pol_benf_end_date_new,
                str_pol_benf_prem_stop_date_new,
                str_policy_pol_ref_no)
    SQL_pol_benf_17     = "UPDATE pol_benf\n   SET PAYMENT_TERM = {},\n       TERM = {},\n       END_DATE = '{}',\n       PREM_STOP_DATE = '{}'\n WHERE POL_REF_NO = {}\n   AND BENFNO = 17;"\
        .format(str_pol_benf_payment_term_new,
                str_pol_benf_term_new,
                str_pol_benf_end_date_new,
                str_pol_benf_prem_stop_date_17_new,
                str_policy_pol_ref_no)
    SQL_pol_benf_act   = list()
    SQL_pol_endorsments = list()
    if data["benfnos"]:
        for benfno in data["benfnos"]:
            SQL_pol_benf_act.append("UPDATE pol_benf\n   SET BENF_STAT = 'A'\n WHERE BENFNO = {};".format(benfno))
            SQL_pol_endorsments.append("INSERT\n  INTO pol_endorsements\n       (SITENO,\n        ENDORSE_TYPE,\n        POL_REF_NO,\n        BENFNO\n        CHANGE_DESC,\n        OLD_VALUE,\n        NEW_VALUE,\n        EFFECTIVE_DATE,\n        TRANSACTION_DATE,\n        LETTER_SENT,\n        STATUS)\nVALUES ({},\n        {},\n        {},\n        {},\n        '{}',\n        '{}',\n        '{}',\n        '{}',\n        {},\n        '{}',\n        '{}');"\
                .format(7,
                        21,
                        str_policy_pol_ref_no,
                        benfno,
                        "Reinstate Benefits",
                        "Lapsed",
                        "Active",
                        str_policy_payout_start_date_minus_1,
                        "sysdate",
                        "N",
                        "A"))
    SQL_pol_endorsments.append("INSERT\n  INTO pol_endorsements\n       (SITENO,\n        ENDORSE_TYPE,\n        POL_REF_NO,\n        CHANGE_DESC,\n        OLD_VALUE,\n        NEW_VALUE,\n        EFFECTIVE_DATE,\n        TRANSACTION_DATE,\n        LETTER_SENT,\n        STATUS)\nVALUES ({},\n        {},\n        {},\n        '{}',\n        {},\n        {},\n        '{}',\n        {},\n        '{}',\n        '{}');"\
        .format(7,
                47,
                str_policy_pol_ref_no,
                "Podaljšanje dobe varčevanja",
                str_policy_term,
                str_policy_term_new,
                sign_date,
                "sysdate",
                "N",
                "A"))
    if data["scholar_rep"]:
        SQL_scholar_rep = "DELETE\n  FROM scholar_rep\n WHERE POLICY_NO = '{}';"\
            .format(str_policy_policy_no)
        SQL_claim_pers_dets = "UPDATE claim_pers_dets\n   SET SET_CODE = 'R'\n WHERE CLAIM_ID = {};"\
            .format(data["claim_pers_dets_claim_id"])
        SQL_sa_postings = "DELETE\n  FROM sa_postings\n WHERE SA_POSTINGNO IN {};".format(tuple(data["sa_postings_sa_postingno"]))
        SQL_claim_dets_7 = "INSERT\n  INTO claim_dets\n       (POL_REF_NO,\n        POLICY_NO,\n        CLAIM_ID,\n        CLAIM_AMT,\n        CLAIM_STAGE,\n        DESCRIPTION,\n        BENFNO)\nVALUES ({},\n        '{}',\n        {},\n        {},\n        {},\n        '{}',\n        {});"\
            .format(str_policy_pol_ref_no,
                    str_policy_policy_no,
                    data["claim_dets_claims"]["claim_stage_1"]["claim_id"],
                    0,
                    7,
                    "Cancelled Surrender (Partial) Claim",
                    data["claim_dets_claims"]["claim_stage_1"]["benfno"])
        SQL_claim_dets_11 = "INSERT\n  INTO claim_dets\n       (POL_REF_NO,\n        POLICY_NO,\n        CLAIM_ID,\n        CLAIM_AMT,\n        CLAIM_STAGE,\n        DESCRIPTION,\n        BENFNO)\nVALUES ({},\n        '{}',\n        {},\n        {},\n        {},\n        '{}',\n        {});"\
            .format(str_policy_pol_ref_no,
                    str_policy_policy_no,
                    data["claim_dets_claims"]["claim_stage_5"]["claim_id"],
                    0,
                    11,
                    "Reinstate Partial Surrender",
                    data["claim_dets_claims"]["claim_stage_5"]["benfno"])
        SQL_ua_holdings = list()
        for fund in data["ua_holdings_all"].keys():
            SQL_ua_holdings_174 = "INSERT\n  INTO ua_holdings\n       (SITENO,\n        DUE_DATE,\n        EFF_DATE,\n        POL_REF_NO,\n        FUNDNO,\n        CURR_NO,\n        BASIC_PREM,\n        TOTAL_PREM,\n        NET_PREM,\n        UNIT_PREM,\n        UNIT_PRICE,\n        UNITS_ALLOCATED,\n        UA_TRANS,\n        BONUS_APP,\n        BENFNO,\n        BENF_ORD,\n        SAV_BENFNO,\n        SAV_BENF_ORD,\n        UA_TRANSNO,\n        UA_EVENTNO)\nVALUES ({},\n        '{}',\n        '{}',\n        {},\n        {},\n        {},\n        {},\n        {},\n        {},\n        {},\n        {},\n        {},\n        '{}',\n        '{}',\n        {},\n        {},\n        {},\n        {},\n        {},\n        {});"\
                .format(7,
                        data["ua_holdings_all"][fund][173]["due_date"].strftime("%#d.%#m.%Y"),
                        data["ua_holdings_all"][fund][173]["eff_date"].strftime("%#d.%#m.%Y"),
                        str_policy_pol_ref_no,
                        fund,
                        1,
                        round(data["ua_holdings_all"][fund][173]["basic_prem"] * -1, 2),
                        round(data["ua_holdings_all"][fund][173]["total_prem"] * -1, 2),
                        round(data["ua_holdings_all"][fund][173]["net_prem"] * -1, 2),
                        round(data["ua_holdings_all"][fund][173]["unit_prem"] * -1, 2),
                        data["ua_holdings_all"][fund][173]["unit_price"],
                        data["ua_holdings_all"][fund][173]["units_allocated"] * -1,
                        "E",
                        "N",
                        data["ua_holdings_all"][fund][173]["benfno"],
                        1,
                        17,
                        1,
                        data["ua_holdings_all"][fund][173]["ua_transno"],
                        174)
            SQL_ua_holdings_176 = "INSERT\n  INTO ua_holdings\n       (SITENO,\n        DUE_DATE,\n        EFF_DATE,\n        POL_REF_NO,\n        FUNDNO,\n        CURR_NO,\n        BASIC_PREM,\n        TOTAL_PREM,\n        NET_PREM,\n        UNIT_PREM,\n        UNIT_PRICE,\n        UNITS_ALLOCATED,\n        UA_TRANS,\n        BONUS_APP,\n        BENFNO,\n        BENF_ORD,\n        SAV_BENFNO,\n        SAV_BENF_ORD,\n        UA_TRANSNO,\n        UA_EVENTNO)\nVALUES ({},\n        '{}',\n        '{}',\n        {},\n        {},\n        {},\n        {},\n        {},\n        {},\n        {},\n        {},\n        {},\n        '{}',\n        '{}',\n        {},\n        {},\n        {},\n        {},\n        {},\n        {});" \
                .format(7,
                        data["ua_holdings_all"][fund][175]["due_date"].strftime("%#d.%#m.%Y"),
                        data["ua_holdings_all"][fund][175]["eff_date"].strftime("%#d.%#m.%Y"),
                        str_policy_pol_ref_no,
                        fund,
                        1,
                        round(data["ua_holdings_all"][fund][175]["basic_prem"] * -1, 2),
                        round(data["ua_holdings_all"][fund][175]["total_prem"] * -1, 2),
                        round(data["ua_holdings_all"][fund][175]["net_prem"] * -1, 2),
                        round(data["ua_holdings_all"][fund][175]["unit_prem"] * -1, 2),
                        data["ua_holdings_all"][fund][175]["unit_price"],
                        data["ua_holdings_all"][fund][175]["units_allocated"] * -1,
                        "E",
                        "N",
                        data["ua_holdings_all"][fund][175]["benfno"],
                        1,
                        17,
                        1,
                        data["ua_holdings_all"][fund][175]["ua_transno"],
                        176)
            SQL_ua_holdings.append(SQL_ua_holdings_174)
            SQL_ua_holdings.append(SQL_ua_holdings_176)


        return "<pre>\n<code class=\"sql\">\n{}\n\n{}\n\n{}\n\n{}\n\n{}\n\n{}\n\n{}\n\n{}\n\n{}\n\n{}\n\n{}\n</code>\n</pre>\n\nPovzetek popravkov iz zgornjih SQL ukazov:\n\nPriložen izvoz trenutnega stanja tabel za polico {}." \
            .format(SQL_policy,
                    SQL_pol_benf,
                    SQL_pol_benf_17,
                    "\n\n".join(SQL_pol_benf_act),
                    SQL_scholar_rep,
                    SQL_claim_pers_dets,
                    SQL_sa_postings,
                    SQL_claim_dets_7,
                    SQL_claim_dets_11,
                    "\n\n".join(SQL_ua_holdings),
                    "\n\n".join(SQL_pol_endorsments),
                    data["policy_policy_no"])
    else:
        return "<pre>\n<code class=\"sql\">\n{}\n\n{}\n\n{}\n\n{}\n</code>\n</pre>\n\nPovzetek popravkov iz zgornjih SQL ukazov:\n\nPriložen izvoz trenutnega stanja tabel za polico {}." \
            .format(SQL_policy,
                    SQL_pol_benf,
                    SQL_pol_benf_17,
                    SQL_pol_endorsments,
                    data["policy_policy_no"])

def main():
    path = getcwd().replace("\\", "\\\\") + "\\\\"
    while True:
        filename_input = input("Ime .xlsx datoteke s tabelami o polici (X + ENTER za izhod): ")
        if filename_input in ("x", "X"): quit()
        try:
            if filename_input.find(".") == -1: filename_input += ".xlsx"
            wb = getWorkbook(filename_input, path)
            break
        except FileNotFoundError: print("Datoteka ne obstaja. Poskusi znova.")
    while True:
        extension_input = input("Doba podaljšanja v mesecih: ")
        if extension_input in ("x", "X"): quit()
        try:
            extension = int(extension_input)
            break
        except: print("Vnesi veljavno dobo podaljšanja.")
    years = extension // 12
    months = extension % 12
    while True:
        sign_date = input("Datum podpisa v formatu D.M.YYYY: ")
        if sign_date in ("x", "X"): quit()
        if verifyDate(sign_date): break
        else: print("Vnesi datum v veljavnem D.M.YYYY formatu.")
    fh = open("podaljsanje_dobe.txt", "w")
    fh.write(generateOutput(getData(wb), years, months, sign_date))
    fh.close()
    print("Ustvarjeni podatki shranjeni v datoteko podaljsanje_dobe.txt.")

if __name__ == "__main__":
    main()