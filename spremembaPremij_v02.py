"""
Script for automatic processing of "Sprememba premij/nevarnostnega razreda" request tickets.
Requires exported tables e2_nezgoda_premija and e2_nezgodna_premija_bruto for a given policy in a .xlsx format. Use 'SELECT * FROM e2_nezgoda_premija WHERE sifra_ponudbe = <policy_code>; SELECT * FROM e2_nezgoda_premija_bruto WHERE sifra_ponudbe = <policy_code>;' SQL commands to generate data for export.
Generates appropriate INSERT SQL commands.
Instructions for use:
    1. export the tables e2_nezgoda_premija and e2_nezgoda_premija_bruto for a given policy and save data as a single .xlsx file in the directory containing this script
    3. run spremembaPremijNevarnostnegaRazreda.py
    4. enter file name of .xlsx file
    5. Enter effective date for new premium values
    6. Enter new monthly and annual premium values for given benefits if applicable.
    5. generated data is saved to sprememba_premij.txt file in the directory containing this script
"""
from os import getcwd
from openpyxl import load_workbook
from collections import OrderedDict
from typing import Iterable

def verifyDate(date_string):
    """
    Verifies input date in D.M.YYYY format.
    :param date_string: str; string of date in D.M.YYYY format
    :return: True if valid date, False otherwise
    """
    if len(date_string) < 8 or len(date_string) > 10: return False
    for i in range(len(date_string)):
        if date_string[1] == ".":
            if int(date_string[0]) <= 0: return False
            if date_string[3] == ".":
                if int(date_string[2]) <= 0: return False
                if int(date_string[4:]) <= 999: return False
            else:
                if int(date_string[2:4]) <= 0 or int(date_string[2:4]) > 12: return False
                if int(date_string[5:]) <= 999: return False
        else:
            if int(date_string[:2]) <= 0 or int(date_string[:2]) > 31: return False
            if date_string[4] == ".":
                if int(date_string[3]) <= 0: return False
                if int(date_string[5:]) <= 999: return False
            else:
                if int(date_string[3:5]) <= 0 or int(date_string[3:5]) > 12: return False
                if int(date_string[6:]) <= 999: return False
    return True

def getWorkbook(filename, dir):
    """
    Fetches Excel Workbook object from a given .xlsx filename.
    :param filename: str; .xlsx filename
    :param dir: str; current working directory path
    :return: Workbook object
    """
    return load_workbook(dir + filename, data_only = True)

def getPolicyNo(workbook):
    """
    Returns policy number from the table.
    :param workbook: Workbook object
    :return: int; policy number
    """
    try: sheet = workbook["Select e2_nezgoda_premija"]
    except: return -1
    return int(sheet["B"][2].value)

def getBenefitIDs(dataset:tuple) -> tuple:
    """
    Returns benefit id numbers from dataset.
    :param data: tuple; dataset for policy query result
    :return: tuple of ints; id number of benefits on policy
    """
    benefit_ids = set()
    index_benfno:int = None
    for i in range(len(dataset[0])):
        if dataset[0][i] == "ID_KRITJE":
            index_benfno = i
            break
    if index_benfno == None:
        print("Med podatki ni stolpca ID_KRITJE.")
        quit()
    for i in range(len(dataset[1])):
        benefit_ids.add(dataset[1][i][index_benfno])
    return tuple(sorted(list(benefit_ids)))

def getBenefit(benefit_id):
    """
    Returns description pertaining to benefit_id.
    :param benefit_id: int; id number of benefit
    :return: str; description of benefit
    """
    benefits = {
        18: 'Smrt',
        63: 'Nezgodna smrt',
        64: 'Nezgodna smrt v prometni nesreči',
        65: 'Trajna invalidnost',
        66: 'Nadomestilo za aktivno zdravljenje',
        67: 'Bolnišnični dan',
        68: 'Nadomestilo za zlom kosti',
         3: 'Kritične bolezni',
        10: 'Nezgodna smrt oz. popolna trajna invalidnost',
        20: 'Nezgodna smrt',
        21: 'Nezgodna popolna trajna invalidnost',
        50: 'Senior',
        60: 'INZ',
        11: 'Smrt',
         5: 'Smrt',
        14: 'Nezgodna smrt oz. popolna trajna invalidnost (odrasli)',
        15: 'Smrt (dodatno)',
        16: 'Nezgodna smrt oz. popolna trajna invalidnost (upravičenec za štipendijo)',
        19: 'Kolektivno ŽZK',
         4: 'Rojstvo otroka',
        70: 'Kolektivno ŽZL',
        61: 'Nezgodna smrt (PK in OR)',
        62: 'Nezgodna popolna trajna invalidnost (PK in OR)',
         1: 'Unit Linked Regular Investment',
         6: 'Unit Linked Single Investment',
         7: 'Universal Life Regular Investment',
         8: 'Universal Life Single Investment',
        98: 'AdHoc Premium',
        12: 'Capital Protector Investment',
        17: 'NV Prva varčevanje',
        13: 'Smrt',
         2: 'Nezgodna smrt',
         9: 'Odgovorna',
        35: 'Varna',
        69: 'Smrt zaradi bolezni',
        71: 'Stroški pogreba',
        72: 'Nezgodna renta',
        73: 'Nezgodni travmatični dogodki',
        74: 'Nezgodna premija',
        75: 'Nadomestilo za invalidnost v prometni nesreči',
        76: 'Nadomestilo za najtežje poškodbe',
        33: 'Kritje 5 bolezni',
        36: 'Smrt',
        77: 'Nadomestilo za okrevanje po poškodbah',
        78: 'Nadomestilo za fizioterapije',
        22: 'Smrt',
        23: 'Smrt'
    }
    return benefits[benefit_id]

def getChanges(benefit_ids:Iterable) -> tuple:
    """
    Returns changes of values for premiums, if applicable
    :param benefit_ids: tuple of ints; id number of benefits on policy
    :return: dict; dictionary of changes of values for premiums
    """
    changes = dict()
    for benefit_id in benefit_ids:
        premium_changes = list()
        while True:
            change = input("Nova vrednost MESEČNE premije za kritje '{}' (ID = {}) oziroma prazno, če ni spremembe: "\
                           .format(getBenefit(benefit_id),
                                   benefit_id))
            if change == "": break
            else:
                try:
                    change = float(change.strip().replace(",", "."))
                    if change: premium_changes.append(change)
                    break
                except: print("Nova vrednost premije mora biti številka ali prazno.")
        if premium_changes:
            while True:
                change = input("Nova vrednost LETNE premije za kritje '{}' (ID = {}): "\
                               .format(getBenefit(benefit_id),
                                       benefit_id))
                try:
                    change = float(change.strip().replace(",", "."))
                    if change: premium_changes.append(change)
                    break
                except: print("Ker je bila spremenjena MESEČNA premija za kritje '{}' (ID = {}), mora tudi nova vrednost LETNE premije biti številka."\
                              .format(getBenefit(benefit_id),
                                     benefit_id))
        if premium_changes: changes[benefit_id] = tuple(premium_changes)
    premium_changes_bruto = list()
    while True:
        change = input("Nova vrednost MESEČNE bruto premije oziroma prazno, če ni spremembe: ")
        if change == "":
            break
        else:
            try:
                change = float(change.strip().replace(",", "."))
                if change: premium_changes_bruto.append(change)
                break
            except:
                print("Nova vrednost premije mora biti številka ali prazno.")
    if premium_changes_bruto:
        while True:
            change = input("Nova vrednost LETNE bruto premije: ")
            try:
                change = float(change.strip().replace(",", "."))
                if change: premium_changes_bruto.append(change)
                break
            except: print("Ker je bila spremenjena MESEČNA bruto premija, mora tudi nova vrednost LETNE bruto premije biti številka.")
    if premium_changes_bruto: changes["bruto"] = tuple(premium_changes_bruto)
    return changes

def getPolicyDataDB(db:str, username:str, password:str, queries:Iterable, variables:Iterable) -> tuple:
    """
    Returns existing data from DB.
    :param db: str; database name
    :param username: str; username for database access
    :param password: str; password for database access
    :param queries: Iterable; an array of SQL queries
    :param variables: Iterable; variables to plug into query (length of input variables must match length of query variables)
    :param policy: str; policy_no to fetch data for
    :return: tuple; results from query/queries
    """
    import DButil as DBu
    for query in queries: assert len(query[1]) == len(variables)
    results:list = list()
    for query in queries:
        for i in range(len(query[1])): results.append(DBu.getQueryResults(db, username, password, query[0], {query[1][i]: variables[i]}))
    return results

def generateOutput(policy_code, data, date):
    """
    Generates output data for 'Sprememba premij' from given input.
    :param policy_code: int; policy code number
    :param data: dict; input data
    :return: str; generated output data
    """
    description1 = "V tabeli e2_nezgoda_premija in e2_nezgoda_premija_bruto je treba za polico {} vnesti nove zapise z datumom veljavnosti {}:\n\n"\
                   .format(policy_code,
                           date)
    descrption2 = "\n\nPriložen izvoz trenutnega stanja tabel e2_nezgoda_premija in e2_nezgoda_premija_bruto za polico {}."\
                  .format(policy_code)
    SQL_e2_nezgoda_premija_insert = ""
    for (id, premiums) in data.items():
        if id != "bruto":
            SQL_e2_nezgoda_premija_insert += "INSERT\n  INTO e2_nezgoda_premija\n      (SIFRA_PONUDBE,\n       ID_KRITJE,\n       MESECNA_PREMIJA,\n       LETNA_PREMIJA,\n       DATUM_VELJAVNOSTI)\nVALUES ('{}',\n       {},\n       {},\n       {},\n       '{}');\n\n"\
                .format(policy_code,
                        id,
                        premiums[0],
                        premiums[1],
                        date)
    SQL_e2_nezgoda_premija_insert += "INSERT\n  INTO e2_nezgoda_premija_bruto\n      (SIFRA_PONUDBE,\n       DATUM_SPREMEMBE,\n       MESECNA_PREMIJA,\n       LETNA_PREMIJA,\n       STATUS)\nVALUES ('{}',\n       '{}',\n       {},\n       {},\n       '{}');\n\n" \
        .format(policy_code,
                date,
                data["bruto"][0],
                data["bruto"][1],
                'A')
    return description1 + "<pre>\n<code class=\"sql\">\n" + SQL_e2_nezgoda_premija_insert[:-1] + "</code>\n</pre>" + descrption2

def main():
    from XLSXutil import saveToXLSX
    from datetime import datetime
    path:str = getcwd().replace("\\", "\\\\") + "\\\\"
    queries:tuple = (("SELECT * FROM e2_nezgoda_premija WHERE sifra_ponudbe = :policy_no", ("policy_no",)), ("SELECT * FROM e2_nezgoda_premija_bruto WHERE sifra_ponudbe = :policy_no", ("policy_no",)))
    today:str = datetime.today().strftime('%Y-%m-%d')
    claim:str = input("Zahtevek: ")
    if not claim.startswith("#"): claim = "#" + claim
    policy_no:str = input("Številka police (X + ENTER za izhod): ")
    if policy_no in ("x", "X"): quit()
    db_username:str = input("Uporabniško ime za dostop do baze PROD19c (X + ENTER za izhod): ")
    if db_username in ("x", "X"): quit()
    db_password:str = input("Geslo za dostop do baze PROD19c (X + ENTER za izhod): ")
    if db_password in ("x", "X"): quit()
    while True:
        effective_date:str = input("Datum podpisa/veljavnosti v formatu D.M.YYYY: ")
        if effective_date in ("x", "X"): quit()
        if verifyDate(effective_date): break
        else: print("Vnesi datum v veljavnem D.M.YYYY formatu.")
    data:tuple = getPolicyDataDB("PROD19c", db_username, db_password, queries, (policy_no,))
    fh = open(f"{claim} {policy_no} sprememba premij SQL {today}.txt", "w")
    fh.write(generateOutput(policy_no, OrderedDict(getChanges(getBenefitIDs(data[0]))), effective_date))
    fh.close()
    saveToXLSX(path, f"{claim} {policy_no} backup {today}", data, ("e2_nezgoda_premija", "e2_nezgoda_premija_bruto"))
    print(f"\nBackup podatki shranjeni v datoteko '{claim} {policy_no} backup {today}.xlsx'.")
    print(f"SQL ukazi shranjeni v datoteko '{claim} {policy_no} sprememba premij SQL {today}.txt'.")

if __name__ == "__main__":
    main()